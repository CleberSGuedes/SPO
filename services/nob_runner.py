from __future__ import annotations

import json
import re
import shutil
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import event, text
from sqlalchemy.exc import SQLAlchemyError

from models import db
from services.job_status import read_cancel_flag, update_status_fields

BATCH_SIZE = 1000
INPUT_DIR = Path("upload/nob")
OUTPUT_DIR = Path("outputs/td_nob")
SAVE_NOB_SHEET = False

COLUMN_ALIASES = {
    "N¶§ NOB": "Nº NOB",
    "N¶§ NOB Estorno/Estornado": "Nº NOB Estorno/Estornado",
    "N¶§ LIQ": "Nº LIQ",
    "N¶§ EMP": "Nº EMP",
    "N¶§ PED": "Nº PED",
    "ExercÇðcio": "Exercício",
    "DotaÇõÇœo OrÇõamentÇ­ria": "Dotação Orçamentária",
    "DevoluÇõÇœo GCV": "Devolução GCV",
    "HistÇürico LIQ": "Histórico LIQ",
    "Data CrÇ¸dito": "Data Crédito",
    "Tipo de TransmissÇœo": "Tipo de Transmissão",
    "TransmissÇœo": "Transmissão",
    "SituaÇõÇœo": "Situação",
    "SituaÇõÇœo OBN": "Situação OBN",
    "SituaÇõÇœo Funcional": "Situação Funcional",
    "Tipo de VÇðnculo": "Tipo de Vínculo",
    "N¶§ ConvÇ¦nio (Ingresso)": "Nº Convênio (Ingresso)",
    "N¶§ ConvÇ¦nio (Repasse)": "Nº Convênio (Repasse)",
    "N¶§ Folha": "Nº Folha",
    "N¶§ PAC": "Nº PAC",
    "N¶§ NOBLIST": "Nº NOBLIST",
    "N¶§ RE OBN": "Nº RE OBN",
    "N¶§ Lote OBN": "Nº Lote OBN",
    "N¶§ Retorno OBN": "Nº Retorno OBN",
    "N¶§ NEX(s)": "Nº NEX(s)",
    "N¶§ Proc Orc Pagto": "Nº Proc Orc Pagto",
    "N¶§ Proc Fin Pagto": "Nº Proc Fin Pagto",
    "Tipo Pagto": "Tipo Pagto",
    "N¶§ C/C+DV (DÇ¸bito)": "Nº C/C+DV (Débito)",
    "N¶§ C/C+DV (CrÇ¸dito)": "Nº C/C+DV (Crédito)",
    "N¶§ Conta STR (CrÇ¸dito)": "Nº Conta STR (Crédito)",
    "DEPJU (CrÇ¸dito)": "DEPJU (Crédito)",
    "Identificador DEPJU (CrÇ¸dito)": "Identificador DEPJU (Crédito)",
    "Banco (DÇ¸bito)": "Banco (Débito)",
    "Banco (CrÇ¸dito)": "Banco (Crédito)",
    "AgÇ¦ncia (DÇ¸bito)": "Agência (Débito)",
    "AgÇ¦ncia (CrÇ¸dito)": "Agência (Crédito)",
    "PerÇðodo/CompetÇ¦ncia": "Período/Competência",
    "ExercÇðcio da Folha": "Exercício da Folha",
    "MÇ¦s da Folha": "Mês da Folha",
    "DescriÇõÇœo do Tipo de OB": "Descrição do Tipo de OB",
    "Data de OcorrÇ¦ncia": "Data de Ocorrência",
    "ObservaÇõÇœo OBN": "Observação OBN",
    "NÇ§mero LicitaÇõÇœo": "Número Licitação",
    "N¶§ ReferÇ¦ncia": "Nº Referência",
    "N¶§ AutenticaÇõÇœo BancÇ­ria": "Nº Autenticação Bancária",
    "CPF/CNPJ do Credor (PremiaÇõÇœo - Nota MT)": "CPF/CNPJ do Credor (Premiação - Nota MT)",
    "Nome do Credor (PremiaÇõÇœo - Nota MT)": "Nome do Credor (Premiação - Nota MT)",
    "N¶ø Proc Judicial RPV": "Nº Proc Judicial RPV",
    "Quantidade de Dias/EficiÇ¦ncia": "Quantidade de Dias/Eficiência",
    "NÇðvel de EficiÇ¦ncia": "Nível de Eficiência",
    "Justificativa (AlteraÇõÇœo da Ordem CronolÇügica)": "Justificativa (Alteração da Ordem Cronológica)",
    "N¶§ Emenda (EP)": "Nº Emenda (EP)",
    "N¶§ ABJ": "Nº ABJ",
    "N¶§ Proc Sequestro Judicial": "Nº Proc Sequestro Judicial",
    "N¶§ RDR": "Nº RDR",
    "N¶§ RDE": "Nº RDE",
    "N¶§ CAD": "Nº CAD",
    "N¶§ CartÇœo Pagto Governo": "Nº Cartão Pagto Governo",
    "N¶§ Conta CartÇœo": "Nº Conta Cartão",
    "SituaÇõÇœo VIPF": "Situação VIPF",
    "N¶§ Lote VIPF": "Nº Lote VIPF",
    "Data OcorrÇ¦ncia VIPF": "Data Ocorrência VIPF",
    "N¶§ Arquivo Retorno VIPF": "Nº Arquivo Retorno VIPF",
    "ObservaÇõÇœo VIPF": "Observação VIPF",
    "Tipo de Transmissao": "Tipo de Transmissao",
    "Data DÇ¸bito": "Data Débito",
    "UsuÇ­rio (InclusÇœo)": "Usuário (Inclusão)",
    "N¶§ DAR Virtual (Registros IntraorÇõamentÇ­rios)": "Nº DAR Virtual (Registros Intraorçamentários)",
    "FunÇõÇœo": "Função",
    "SubfunÇõÇœo": "Subfunção",
}

COLUMNS_TO_DROP = [
    "UO Extinta",
    "Nº Proc Orc Pagto",
    "Nº Proc Fin Pagto",
    "Nº PAC",
    "Nº NOBLIST",
    "Tipo Pagto",
    "CBA",
    "Natureza",
    "Banco (Débito)",
    "Agência (Débito)",
    "Nº C/C+DV (Débito)",
    "Subconta",
    "Forma de Recebimento",
    "Banco (Crédito)",
    "Agência (Crédito)",
    "Nº C/C+DV (Crédito)",
    "Nº Conta STR (Crédito)",
    "DEPJU (Crédito)",
    "Identificador DEPJU (Crédito)",
    "Nome do Ordenador de Despesa",
    "Nome do Liberador de Pagamento",
    "Situação",
    "REG",
    "Período/Competência",
    "UO SEAP",
    "Exe Anterior Folha",
    "Exercício da Folha",
    "Mês da Folha",
    "Tipo de Folha",
    "Nº Folha",
    "Situação Funcional",
    "Tipo de Vínculo",
    "Indicativo de NOB/Fatura Fato 54",
    "Tipo de Transmissão",
    "Transmissão",
    "Situação OBN",
    "Tipo OB",
    "Descrição do Tipo de OB",
    "Nº RE OBN",
    "Nº Lote OBN",
    "Data de Ocorrência",
    "Nº Retorno OBN",
    "Retorno OBN",
    "Retorno CNAB240",
    "Observação OBN",
    "Data Crédito",
    "Nº NEX(s)",
    "Valor LIQ",
    "Nº Convênio (Ingresso)",
    "Nº Convênio (Repasse)",
    "Modalidade de Licitação",
    "Número Licitação",
    "Ano da Licitação",
    "Fundamento Legal",
    "Entrega Imediata",
    "NEX/NOB/OBF na RE",
    "Tipo de Fatura",
    "Subtipo de Fatura",
    "Tributo Federal",
    "Nº Referência",
    "Valor da Fatura",
    "Valor da Multa",
    "Valor dos Juros/Encargos",
    "Nº Autenticação Bancária",
    "CPF/CNPJ do Credor (Premiação - Nota MT)",
    "Nome do Credor (Premiação - Nota MT)",
    "Nº Proc Judicial RPV",
    "Nº ABJ",
    "Nº Proc Sequestro Judicial",
    "Nº Emenda (EP)",
    "Autor da Emenda (EP)",
    "Quantidade de Dias/Eficiência",
    "Nível de Eficiência",
    "Justificativa (Alteração da Ordem Cronológica)",
    "Nº RDR",
    "Nº RDE",
    "Nº CAD",
    "Nº Cartão Pagto Governo",
    "Nº Conta Cartão",
    "Situação VIPF",
    "Nº Lote VIPF",
    "Data Ocorrência VIPF",
    "Nº Arquivo Retorno VIPF",
    "Data Retorno VIPF",
    "Cod Retorno VIPF",
    "Observação VIPF",
    "Tipo de Transmissao",
    "Data Débito",
    "Usuário (Inclusão)",
    "Nº DAR Virtual (Registros Intraorçamentários)",
]

REQUIRED_COLS_RAW = list(
    {
        "N¶§ NOB",
        "Nº NOB",
        "N¶§ NOB Estorno/Estornado",
        "Nº NOB Estorno/Estornado",
        "N¶§ LIQ",
        "Nº LIQ",
        "N¶§ EMP",
        "Nº EMP",
        "N¶§ PED",
        "Nº PED",
        "Valor NOB",
        "DevoluÇõÇœo GCV",
        "Devolução GCV",
        "Data NOB",
        "Data Cadastro NOB",
        "Data/Hora de Cadastro da LIQ",
        "DotaÇõÇœo OrÇõamentÇ­ria",
        "Dotação Orçamentária",
        "Natureza de Despesa",
        "Nome da Fonte de Recurso",
        "ExercÇðcio",
        "Exercício",
        "UG",
        "UO",
        "Nome do Credor Principal",
        "CPF/CNPJ do Credor Principal",
        "Credor",
        "Nome do Credor",
        "CPF/CNPJ do Credor",
        "HistÇürico LIQ",
        "Histórico LIQ",
    }
)

COL_MAP = {
    "exercicio": "exercicio",
    "n_nob": "numero_nob",
    "no_nob": "numero_nob",
    "n_nob_estorno_estornado": "numero_nob_estorno",
    "no_nob_estorno_estornado": "numero_nob_estorno",
    "n_liq": "numero_liq",
    "no_liq": "numero_liq",
    "n_emp": "numero_emp",
    "no_emp": "numero_emp",
    "n_ped": "numero_ped",
    "no_ped": "numero_ped",
    "valor_nob": "valor_nob",
    "devolucao_gcv": "devolucao_gcv",
    "valor_nob_gcv": "valor_nob_gcv",
    "data_nob": "data_nob",
    "data_cadastro_nob": "data_cadastro_nob",
    "data_hora_de_cadastro_da_liq": "data_hora_cadastro_liq",
    "dotacao_orcamentaria": "dotacao_orcamentaria",
    "natureza_de_despesa": "natureza_despesa",
    "nome_da_fonte_de_recurso": "nome_fonte_recurso",
    "ug": "ug",
    "uo": "uo",
    "nome_do_credor_principal": "nome_credor_principal",
    "cpf_cnpj_do_credor_principal": "cpf_cnpj_credor_principal",
    "credor": "credor",
    "nome_do_credor": "nome_credor",
    "cpf_cnpj_do_credor": "cpf_cnpj_credor",
    "historico_liq": "historico_liq",
    "empenho_atual": "empenho_atual",
    "empenho_rp": "empenho_rp",
    "funcao": "funcao",
    "subfuncao": "subfuncao",
    "programa_de_governo": "programa_governo",
    "paoe": "paoe",
    "cat_econ": "cat_econ",
    "grupo": "grupo",
    "modalidade": "modalidade",
    "iduso": "iduso",
}

INSERT_COLS = [
    "upload_id",
    "exercicio",
    "numero_nob",
    "numero_nob_estorno",
    "numero_liq",
    "numero_emp",
    "numero_ped",
    "valor_nob",
    "devolucao_gcv",
    "valor_nob_gcv",
    "data_nob",
    "data_cadastro_nob",
    "data_hora_cadastro_liq",
    "dotacao_orcamentaria",
    "natureza_despesa",
    "nome_fonte_recurso",
    "ug",
    "uo",
    "nome_credor_principal",
    "cpf_cnpj_credor_principal",
    "credor",
    "nome_credor",
    "cpf_cnpj_credor",
    "historico_liq",
    "empenho_atual",
    "empenho_rp",
    "funcao",
    "subfuncao",
    "programa_governo",
    "paoe",
    "cat_econ",
    "grupo",
    "modalidade",
    "iduso",
    "raw_payload",
    "data_atualizacao",
    "data_arquivo",
    "user_email",
    "ativo",
]

_FAST_EXEC_ENABLED = False


def ensure_dirs() -> None:
    for base in (INPUT_DIR, OUTPUT_DIR, INPUT_DIR / "tmp", OUTPUT_DIR / "tmp"):
        base.mkdir(parents=True, exist_ok=True)


def move_existing_to_tmp(base_dir: Path) -> None:
    tmp = base_dir / "tmp"
    tmp.mkdir(parents=True, exist_ok=True)
    for f in base_dir.glob("*.xlsx"):
        if f.name.startswith("~$"):
            continue
        dest = tmp / f"{f.stem}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}{f.suffix}"
        try:
            shutil.move(str(f), dest)
        except PermissionError:
            print(f"Aviso: nao foi possivel mover {f} para tmp (arquivo em uso).")


def _normalize_col(name: Any) -> str:
    texto = unicodedata.normalize("NFKD", str(name or ""))
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^\w\s]", " ", texto)
    texto = re.sub(r"\s+", "_", texto).strip("_").lower()
    return texto


def br_to_float_series(s: pd.Series) -> pd.Series:
    s = s.fillna("").astype(str).str.strip()
    vazio = s.eq("") | s.str.upper().eq("NAO INFORMADO") | s.str.upper().eq("NÃO INFORMADO")
    s = s.str.replace("\xa0", " ", regex=False)
    s = s.str.replace("R$", "", regex=False).str.replace(" ", "", regex=False)
    mask_comma = s.str.contains(",", na=False)
    s = s.where(~mask_comma, s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    s = s.str.replace(r"[^0-9.\-]", "", regex=True)
    s = pd.to_numeric(s, errors="coerce").fillna(0.0)
    s = s.where(~vazio, 0.0)
    return s


def format_brl(series: pd.Series) -> pd.Series:
    return series.fillna(0.0).astype(float).map(
        lambda v: f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    )


def normalizar_colunas(df: pd.DataFrame) -> pd.DataFrame:
    renomes = {orig: dest for orig, dest in COLUMN_ALIASES.items() if orig in df.columns}
    if renomes:
        df = df.rename(columns=renomes)
    df.columns = df.columns.str.strip()
    return df


def carregar_planilha_base(xlsx_path: Path) -> pd.DataFrame:
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    cached_rows = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx > 400:
            break
        cached_rows.append([cell for cell in row])

    header_row_idx = None
    header_values = []

    if len(cached_rows) >= 5:
        row5 = cached_rows[4]
        if any(isinstance(c, str) and "exerc" in c.lower() for c in row5):
            header_row_idx = 5
            header_values = [str(c).strip() if c is not None else "" for c in row5]

    if header_row_idx is None:
        candidate_single = None
        for idx, row in enumerate(cached_rows, start=1):
            nonempty = sum(1 for c in row if c not in (None, ""))
            has_exerc = any(isinstance(c, str) and "exerc" in c.lower() for c in row)
            if has_exerc and nonempty >= 3:
                header_row_idx = idx
                header_values = [str(c).strip() if c is not None else "" for c in row]
                break
            if has_exerc and nonempty == 1 and candidate_single is None:
                candidate_single = idx

        if header_row_idx is None and candidate_single:
            next_idx = candidate_single + 1
            if next_idx <= len(cached_rows):
                header_row_idx = next_idx
                header_values = [str(c).strip() if c is not None else "" for c in cached_rows[next_idx - 1]]

    if header_row_idx is None:
        raise ValueError("Cabecalho com 'Exercicio' nao encontrado nas primeiras 400 linhas.")

    def canon(n: str) -> str:
        n = str(n).strip()
        return COLUMN_ALIASES.get(n, n)

    header_norm = [canon(n) for n in header_values]
    required_canon = {canon(c) for c in REQUIRED_COLS_RAW}

    keep_indices = []
    keep_names = []
    for pos, norm_name in enumerate(header_norm):
        if norm_name in required_canon:
            keep_indices.append(pos)
            keep_names.append(norm_name)

    if not keep_indices:
        raise ValueError(f"Nenhuma coluna necessaria encontrada. Header detectado: {header_values}")

    data_rows = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx <= header_row_idx:
            continue
        selected = []
        for pos in keep_indices:
            val = row[pos] if pos < len(row) else None
            selected.append("" if val is None else str(val))
        data_rows.append(selected)

    df = pd.DataFrame(data_rows, columns=keep_names)
    df = normalizar_colunas(df)
    return df


def remover_colunas(df: pd.DataFrame) -> pd.DataFrame:
    return df.drop(columns=[col for col in COLUMNS_TO_DROP if col in df.columns], errors="ignore")


def tratar_colunas_texto(df: pd.DataFrame) -> pd.DataFrame:
    cols = df.select_dtypes(include=["object"]).columns
    if cols.empty:
        return df
    df[cols] = df[cols].astype(str)
    df[cols] = df[cols].replace(r"_x000D_", "", regex=True)
    df[cols] = df[cols].replace(r"[^\S\r\n]+", " ", regex=True)
    df[cols] = df[cols].replace(r"\s+", " ", regex=True)
    df[cols] = df[cols].apply(lambda s: s.str.replace("*", "|"))
    df[cols] = df[cols].replace(
        {"": "NÃO INFORMADO", "nan": "NÃO INFORMADO", "NAO INFORMADO": "NÃO INFORMADO"}
    )
    return df


def tratar_colunas_numericas(df: pd.DataFrame) -> pd.DataFrame:
    col_monetarias = ["Valor NOB", "Devolução GCV"]

    val_nob = br_to_float_series(df[col_monetarias[0]]) if col_monetarias[0] in df.columns else None
    val_gcv = br_to_float_series(df[col_monetarias[1]]) if col_monetarias[1] in df.columns else None

    if val_nob is not None:
        df[col_monetarias[0]] = format_brl(val_nob)
    if val_gcv is not None:
        df[col_monetarias[1]] = format_brl(val_gcv)

    if val_nob is not None and val_gcv is not None:
        df["Valor NOB - GCV"] = format_brl(val_nob - val_gcv)
    else:
        df["Valor NOB - GCV"] = "0,00"

    col_datas = ["Data NOB", "Data Cadastro NOB"]
    for col in col_datas:
        if col in df.columns:
            serie_str = df[col].astype(str).str.strip()
            if serie_str.str.match(r"\d{4}-\d{2}-\d{2}").all():
                seri = pd.to_datetime(serie_str, errors="coerce", dayfirst=False)
            else:
                seri = pd.to_datetime(serie_str, errors="coerce", dayfirst=True)
            df[col] = seri.dt.strftime("%d/%m/%Y").fillna("NÃO INFORMADO")

    return df


def reorganizar_colunas_iniciais(df: pd.DataFrame) -> pd.DataFrame:
    if "Nº NOB" in df.columns:
        index_nob = df.columns.get_loc("Nº NOB") + 1
        colunas_ordem = ["Nº NOB Estorno/Estornado", "Nº LIQ", "Nº EMP", "Nº PED", "Valor NOB"]
        colunas_existentes = [col for col in colunas_ordem if col in df.columns]
        outras_colunas = [col for col in df.columns if col not in colunas_existentes]
        nova_ordem = outras_colunas[:index_nob] + colunas_existentes + outras_colunas[index_nob:]
        return df[nova_ordem]
    return df


def reorganizar_colunas_valor(df: pd.DataFrame) -> pd.DataFrame:
    if "Valor NOB" in df.columns:
        index_valor_nob = df.columns.get_loc("Valor NOB") + 1
        colunas_ordem = ["Devolução GCV", "Valor NOB - GCV"]
        colunas_existentes = [col for col in colunas_ordem if col in df.columns]
        outras_colunas = [col for col in df.columns if col not in colunas_existentes]
        nova_ordem = outras_colunas[:index_valor_nob] + colunas_existentes + outras_colunas[index_valor_nob:]
        return df[nova_ordem]
    return df


def reorganizar_colunas_credor_e_data(df: pd.DataFrame) -> pd.DataFrame:
    df = df.loc[:, ~df.columns.duplicated()]
    if "Data Cadastro NOB" not in df.columns or "Data NOB" not in df.columns:
        print("Erro: 'Data Cadastro NOB' ou 'Data NOB' nao encontrado no DataFrame.")
        return df

    colunas_credor = [
        "Nome do Credor Principal",
        "CPF/CNPJ do Credor Principal",
        "Credor",
        "Nome do Credor",
        "CPF/CNPJ do Credor",
    ]
    colunas_existentes_credor = [col for col in colunas_credor if col in df.columns]

    colunas_lista = list(df.columns)
    colunas_lista.remove("Data NOB")
    nova_posicao_data_nob = colunas_lista.index("Data Cadastro NOB")
    colunas_lista.insert(nova_posicao_data_nob, "Data NOB")
    df_intermediario = df[colunas_lista]

    for col in colunas_existentes_credor:
        colunas_lista.remove(col)
    nova_posicao_data_nob = colunas_lista.index("Data NOB")
    for i, col in enumerate(colunas_existentes_credor):
        colunas_lista.insert(nova_posicao_data_nob + i, col)

    df_final = df_intermediario[colunas_lista]
    return df_final


def adicionar_colunas_empenho(df: pd.DataFrame) -> pd.DataFrame:
    if "Nº EMP" not in df.columns or "Exercício" not in df.columns:
        print("Erro: Coluna 'Nº EMP' ou 'Exercício' nao encontrada!")
        return df

    def extrair_ano(n_emp: str) -> str | None:
        partes = n_emp.split(".")
        if len(partes) >= 3 and partes[2].isdigit():
            return partes[2][-2:]
        return None

    df["Ano Nº EMP"] = df["Nº EMP"].astype(str).apply(extrair_ano)
    df["Ano Exercício"] = df["Exercício"].astype(str).str[-2:]
    df["Empenho Atual"] = df.apply(
        lambda x: x["Nº EMP"] if x["Ano Nº EMP"] == x["Ano Exercício"] else "", axis=1
    )
    df["Empenho RP"] = df.apply(
        lambda x: x["Nº EMP"] if x["Ano Nº EMP"] != x["Ano Exercício"] else "", axis=1
    )
    df["Empenho Atual"] = df["Empenho Atual"].replace("", "NÃO INFORMADO")
    df["Empenho RP"] = df["Empenho RP"].replace("", "NÃO INFORMADO")
    df.drop(columns=["Ano Nº EMP", "Ano Exercício"], inplace=True)

    colunas_lista = list(df.columns)
    index_n_emp = colunas_lista.index("Nº EMP")
    colunas_lista.remove("Empenho Atual")
    colunas_lista.remove("Empenho RP")
    colunas_lista.insert(index_n_emp + 1, "Empenho Atual")
    colunas_lista.insert(index_n_emp + 2, "Empenho RP")
    df = df[colunas_lista]
    return df


def filtrar_nob_estorno(df: pd.DataFrame) -> pd.DataFrame:
    if "Nº NOB Estorno/Estornado" not in df.columns:
        print("Erro: Coluna 'Nº NOB Estorno/Estornado' nao encontrada!")
        return df
    coluna = df["Nº NOB Estorno/Estornado"].astype(str).str.strip().str.upper()
    df_filtrado = df[coluna.isin({"NAO INFORMADO", "NÃO INFORMADO"})].copy()
    print(f"Registros filtrados: {len(df) - len(df_filtrado)} removidos. Restantes: {len(df_filtrado)}")
    return df_filtrado


def adicionar_colunas_dotacao(df: pd.DataFrame) -> pd.DataFrame:
    if "Dotação Orçamentária" not in df.columns:
        print("Erro: Coluna 'Dotação Orçamentária' nao encontrada!")
        return df

    def extrair_elemento(dotacao: Any, posicao: int) -> str:
        partes = str(dotacao).split(".")
        return partes[posicao] if len(partes) > posicao else "NÃO INFORMADO"

    df["Função"] = df["Dotação Orçamentária"].apply(lambda x: extrair_elemento(x, 2))
    df["Subfunção"] = df["Dotação Orçamentária"].apply(lambda x: extrair_elemento(x, 3))
    df["Programa de Governo"] = df["Dotação Orçamentária"].apply(lambda x: extrair_elemento(x, 4))
    df["PAOE"] = df["Dotação Orçamentária"].apply(lambda x: extrair_elemento(x, 5))
    df["Natureza de Despesa"] = df["Dotação Orçamentária"].apply(lambda x: extrair_elemento(x, 7))

    colunas_lista = list(df.columns)
    index_dotacao = colunas_lista.index("Dotação Orçamentária")
    for col in ["Função", "Subfunção", "Programa de Governo", "PAOE", "Natureza de Despesa"]:
        colunas_lista.remove(col)
    for i, col in enumerate(["Função", "Subfunção", "Programa de Governo", "PAOE", "Natureza de Despesa"]):
        colunas_lista.insert(index_dotacao + 1 + i, col)
    df = df[colunas_lista]
    return df


def adicionar_colunas_natureza_despesa(df: pd.DataFrame) -> pd.DataFrame:
    if "Natureza de Despesa" not in df.columns:
        print("Erro: Coluna 'Natureza de Despesa' nao encontrada!")
        return df

    def extrair_elemento(natureza: Any, inicio: int, fim: int) -> str:
        natureza = str(natureza).strip()
        return natureza[inicio:fim] if len(natureza) >= fim else "NÃO INFORMADO"

    df["Cat.Econ"] = df["Natureza de Despesa"].apply(lambda x: extrair_elemento(x, 0, 1))
    df["Grupo"] = df["Natureza de Despesa"].apply(lambda x: extrair_elemento(x, 1, 2))
    df["Modalidade"] = df["Natureza de Despesa"].apply(lambda x: extrair_elemento(x, 2, 4))

    colunas_lista = list(df.columns)
    index_natureza = colunas_lista.index("Natureza de Despesa")
    for col in ["Cat.Econ", "Grupo", "Modalidade"]:
        colunas_lista.remove(col)
    for i, col in enumerate(["Cat.Econ", "Grupo", "Modalidade"]):
        colunas_lista.insert(index_natureza + 1 + i, col)
    df = df[colunas_lista]
    return df


def adicionar_coluna_iduso(df: pd.DataFrame) -> pd.DataFrame:
    if "Dotação Orçamentária" not in df.columns:
        print("Erro: Coluna 'Dotação Orçamentária' nao encontrada!")
        return df

    def extrair_elemento(dotacao: Any, posicao: int) -> str:
        partes = str(dotacao).split(".")
        return partes[posicao] if len(partes) > posicao else "NÃO INFORMADO"

    df["Iduso"] = df["Dotação Orçamentária"].apply(lambda x: extrair_elemento(x, 9))

    colunas_lista = list(df.columns)
    if "Nome da Fonte de Recurso" in colunas_lista:
        index_fonte = colunas_lista.index("Nome da Fonte de Recurso")
    else:
        index_fonte = len(colunas_lista) - 1
    colunas_lista.remove("Iduso")
    colunas_lista.insert(index_fonte + 1, "Iduso")
    df = df[colunas_lista]
    return df


def processar_nob(file_path: Path) -> Path:
    df_nob = carregar_planilha_base(file_path)
    df_limpo = remover_colunas(df_nob.copy())
    df_tratado = tratar_colunas_texto(df_limpo)
    df_tratado = tratar_colunas_numericas(df_tratado)
    df_tratado = reorganizar_colunas_iniciais(df_tratado)
    df_tratado = reorganizar_colunas_valor(df_tratado)
    df_tratado = reorganizar_colunas_credor_e_data(df_tratado)
    df_tratado = adicionar_colunas_empenho(df_tratado)
    df_tratado = filtrar_nob_estorno(df_tratado)
    df_tratado = adicionar_colunas_dotacao(df_tratado)
    df_tratado = adicionar_colunas_natureza_despesa(df_tratado)
    df_tratado = adicionar_coluna_iduso(df_tratado)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_file = OUTPUT_DIR / f"{file_path.stem}_tratado.xlsx"
    writer = pd.ExcelWriter(output_file, engine="xlsxwriter")
    writer.book.strings_to_urls = False
    if SAVE_NOB_SHEET:
        df_nob.to_excel(writer, index=False, sheet_name="nob")
    df_tratado.to_excel(writer, index=False, sheet_name="nob_tratado")
    writer.close()
    print(f"Planilha salva em: {output_file}")
    return output_file


def _clean_val(val: Any) -> Any:
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass
    if isinstance(val, str) and val.strip() == "-":
        return None
    return val


def _parse_valor_db(valor: Any) -> float | None:
    if valor is None:
        return None
    s = str(valor).strip()
    if s in ("", "-", "NAO INFORMADO", "NÃO INFORMADO"):
        return None
    s_num = re.sub(r"[^\d,.-]", "", s)
    if "," in s_num:
        s_num = s_num.replace(".", "").replace(",", ".")
    try:
        return float(s_num)
    except ValueError:
        return None


def _parse_data_db(valor: Any) -> datetime | None:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor
    s = str(valor).strip()
    if not s or s.upper() in ("-", "NAO INFORMADO", "NÃO INFORMADO", "00/00/0000"):
        return None
    s = s.replace("-", "/")
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def montar_registros_para_db(
    df: pd.DataFrame, data_arquivo: datetime, user_email: str, upload_id: int
) -> list[dict[str, Any]]:
    registros: list[dict[str, Any]] = []
    rows = df.to_dict(orient="records")
    for row in rows:
        payload: dict[str, Any] = {}
        for col, val in row.items():
            key = _normalize_col(col)
            db_col = COL_MAP.get(key)
            if not db_col:
                continue
            payload[db_col] = _clean_val(val)

        for col in ("valor_nob", "devolucao_gcv", "valor_nob_gcv"):
            if col in payload:
                payload[col] = _parse_valor_db(payload[col])
        for col in ("data_nob", "data_cadastro_nob", "data_hora_cadastro_liq"):
            if col in payload:
                payload[col] = _parse_data_db(payload[col])

        safe_row: dict[str, Any] = {}
        for k, v in row.items():
            try:
                if pd.isna(v):
                    safe_row[k] = None
                    continue
            except Exception:
                pass
            if hasattr(v, "isoformat"):
                try:
                    safe_row[k] = v.isoformat()
                    continue
                except Exception:
                    pass
            safe_row[k] = v

        payload["raw_payload"] = json.dumps(safe_row, ensure_ascii=False)
        payload["upload_id"] = upload_id
        payload["data_atualizacao"] = datetime.utcnow()
        payload["data_arquivo"] = data_arquivo
        payload["user_email"] = user_email
        payload["ativo"] = True
        for col in INSERT_COLS:
            payload.setdefault(col, None)
        registros.append(payload)
    return registros


def _enable_fast_executemany() -> None:
    global _FAST_EXEC_ENABLED
    if _FAST_EXEC_ENABLED:
        return

    @event.listens_for(db.engine, "before_cursor_execute")
    def _set_fast_exec(conn, cursor, statement, parameters, context, executemany):  # type: ignore[no-redef]
        if executemany:
            try:
                cursor.fast_executemany = True
            except Exception:
                pass

    _FAST_EXEC_ENABLED = True


def update_database(
    df: pd.DataFrame, data_arquivo: datetime, user_email: str, upload_id: int
) -> int:
    insert_sql = text(
        """
        INSERT INTO nob (
            upload_id, exercicio, numero_nob, numero_nob_estorno, numero_liq, numero_emp, numero_ped,
            valor_nob, devolucao_gcv, valor_nob_gcv, data_nob, data_cadastro_nob, data_hora_cadastro_liq,
            dotacao_orcamentaria, natureza_despesa, nome_fonte_recurso, ug, uo, nome_credor_principal,
            cpf_cnpj_credor_principal, credor, nome_credor, cpf_cnpj_credor, historico_liq, empenho_atual,
            empenho_rp, funcao, subfuncao, programa_governo, paoe, cat_econ, grupo, modalidade, iduso,
            raw_payload, data_atualizacao, data_arquivo, user_email, ativo
        )
        VALUES (
            :upload_id, :exercicio, :numero_nob, :numero_nob_estorno, :numero_liq, :numero_emp, :numero_ped,
            :valor_nob, :devolucao_gcv, :valor_nob_gcv, :data_nob, :data_cadastro_nob, :data_hora_cadastro_liq,
            :dotacao_orcamentaria, :natureza_despesa, :nome_fonte_recurso, :ug, :uo, :nome_credor_principal,
            :cpf_cnpj_credor_principal, :credor, :nome_credor, :cpf_cnpj_credor, :historico_liq, :empenho_atual,
            :empenho_rp, :funcao, :subfuncao, :programa_governo, :paoe, :cat_econ, :grupo, :modalidade, :iduso,
            :raw_payload, :data_atualizacao, :data_arquivo, :user_email, :ativo
        )
        """
    )

    try:
        db.session.execute(text("UPDATE nob SET ativo = 0 WHERE ativo = 1"))
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        raise

    _enable_fast_executemany()
    registros = montar_registros_para_db(df, data_arquivo, user_email, upload_id)
    total_registros = len(registros)
    print(f" Gravando {total_registros} registros no banco...")
    update_status_fields(
        "nob",
        upload_id,
        progress=0,
        message=f"Gravando registros no banco (0/{total_registros}).",
    )
    total = 0
    for start in range(0, len(registros), BATCH_SIZE):
        if read_cancel_flag("nob", upload_id):
            raise RuntimeError("PROCESSAMENTO_CANCELADO")
        chunk = registros[start : start + BATCH_SIZE]
        try:
            db.session.execute(insert_sql, chunk)
            db.session.commit()
            total += len(chunk)
            print(f" Inseridos {total}/{total_registros} registros...")
            if total_registros:
                progress = min(100, int((total / total_registros) * 100))
            else:
                progress = 100
            update_status_fields(
                "nob",
                upload_id,
                progress=progress,
                message=f"Gravando registros no banco ({total}/{total_registros}).",
            )
        except SQLAlchemyError:
            db.session.rollback()
            raise
    return total


def run_nob(
    file_path: Path, data_arquivo: datetime, user_email: str, upload_id: int
) -> tuple[int, Path]:
    ensure_dirs()
    move_existing_to_tmp(OUTPUT_DIR)
    output_path = processar_nob(file_path)

    df_tratado = pd.read_excel(output_path, sheet_name="nob_tratado", dtype=str)
    colunas_data = {
        "data_nob",
        "data_cadastro_nob",
        "data_hora_cadastro_liq",
        "data_hora_de_cadastro_da_liq",
        "data_atualizacao",
        "data_arquivo",
    }
    for col in df_tratado.columns:
        if _normalize_col(col) in colunas_data:
            serie_str = df_tratado[col].astype(str).str.strip()
            if serie_str.str.match(r"\d{4}-\d{2}-\d{2}").all():
                df_tratado[col] = pd.to_datetime(serie_str, errors="coerce", dayfirst=False)
            else:
                df_tratado[col] = pd.to_datetime(serie_str, errors="coerce", dayfirst=True)
    total = update_database(df_tratado, data_arquivo, user_email, upload_id)
    return total, output_path
